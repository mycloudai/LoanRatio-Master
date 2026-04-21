"""Excel export using openpyxl with native formulas where feasible.

Layout:
  Sheet "Summary": payers + current ratios + cumulative principal.
  Sheet "Loans": loan list and remaining principal.
  Sheet "Months": one row per month. For auto rows, the interest_share
    and raw_principal cells are written as formulas referring to other
    cells so the user can tweak inputs in Excel.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def export_excel(state: dict[str, Any]) -> bytes:
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Payer ID", "Name", "Cumulative Principal", "Current Ratio"])
    for c in ws[1]:
        c.font = Font(bold=True)
    payers = state.get("payers", [])
    months = state.get("months", [])
    last_per = (months[-1].get("computed", {}).get("perPayer", {}) if months else {})
    for p in payers:
        info = last_per.get(p["id"], {})
        ws.append(
            [
                p["id"],
                p.get("name", ""),
                info.get("cumulativePrincipal", 0.0),
                info.get("ratio", 0.0),
            ]
        )

    # --- Loans ---
    ws_l = wb.create_sheet("Loans")
    ws_l.append(["Loan ID", "Name", "Original Amount", "Remaining Principal"])
    for c in ws_l[1]:
        c.font = Font(bold=True)
    for ln in state.get("loans", []):
        ws_l.append(
            [
                ln["id"],
                ln.get("name", ""),
                float(ln.get("originalAmount", 0.0)),
                float(ln.get("remainingPrincipal", 0.0)),
            ]
        )

    # --- Months ---
    ws_m = wb.create_sheet("Months")
    payer_ids = [p["id"] for p in payers]
    payer_names = {p["id"]: p.get("name", p["id"]) for p in payers}
    header = ["YearMonth", "Mode", "Total Interest"]
    for pid in payer_ids:
        nm = payer_names[pid]
        header += [
            f"{nm} prev_ratio",
            f"{nm} payment",
            f"{nm} interest_share",
            f"{nm} raw_principal",
            f"{nm} adj_principal",
            f"{nm} cum_principal",
            f"{nm} ratio",
        ]
    ws_m.append(header)
    for c in ws_m[1]:
        c.font = Font(bold=True)

    prev_ratio_row = {pid: None for pid in payer_ids}  # cell address per payer

    for m in months:
        per = m.get("computed", {}).get("perPayer", {})
        total_int = m.get("computed", {}).get("totalInterest", 0.0)
        row = [m.get("yearMonth", ""), m.get("mode", "auto"), float(total_int)]
        # row index will be ws_m.max_row + 1 after append
        next_row = ws_m.max_row + 1
        col_idx = 4  # 1-based, after 3 header cols
        # Pre-compute payment lookup
        payments = {pp["payerId"]: float(pp.get("amount", 0.0)) for pp in m.get("payerPayments", []) or []}
        for pid in payer_ids:
            info = per.get(pid, {})
            row += [
                0.0,  # prev_ratio - filled below with formula for auto rows
                payments.get(pid, 0.0),
                info.get("interestShare", 0.0),
                info.get("rawPrincipal", 0.0),
                info.get("adjPrincipal", 0.0),
                info.get("cumulativePrincipal", 0.0),
                info.get("ratio", 0.0),
            ]
            col_idx += 7
        ws_m.append(row)

        # Now overwrite cells with formulas where feasible (only for auto rows).
        is_auto = m.get("mode", "auto") == "auto"
        col_idx = 4
        total_int_cell = f"C{next_row}"
        for pid in payer_ids:
            prev_ratio_col = col_idx
            payment_col = col_idx + 1
            interest_col = col_idx + 2
            raw_col = col_idx + 3
            # prev_ratio cell
            if prev_ratio_row[pid] is not None and is_auto:
                # reference previous month's ratio cell
                prev_addr = prev_ratio_row[pid]
                ws_m.cell(row=next_row, column=prev_ratio_col).value = f"={prev_addr}"
                # interest_share = prev_ratio * total_interest
                ws_m.cell(
                    row=next_row, column=interest_col
                ).value = f"={get_column_letter(prev_ratio_col)}{next_row}*{total_int_cell}"
                # raw_principal = payment - interest_share
                ws_m.cell(
                    row=next_row, column=raw_col
                ).value = (
                    f"={get_column_letter(payment_col)}{next_row}"
                    f"-{get_column_letter(interest_col)}{next_row}"
                )
            # Remember this row's ratio cell for next iteration
            ratio_col = col_idx + 6
            prev_ratio_row[pid] = f"{get_column_letter(ratio_col)}{next_row}"
            col_idx += 7

    # Auto-size-ish: set width 18 for everything
    for sheet in wb.worksheets:
        for i in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(i)].width = 18

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
